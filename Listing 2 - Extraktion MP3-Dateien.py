import os
import shutil
import openpyxl

workbook = openpyxl.load_workbook('GESCHICHTETE_STICHPROBE.xlsx')
sheet = workbook.active

source_folder_path = 'PFAD_QUELLORDNER'
destination_folder_path = 'PFAD_ZIELORDNER'

# Überprüfen, ob der Zielordner existiert, und wenn nicht, wird er erstellt
if not os.path.exists(destination_folder_path):
    os.makedirs(destination_folder_path)

# Durchlaufen aller Zeilen im Arbeitsblatt
for row in range(1, sheet.max_row + 1):
    path = sheet.cell(row=row, column=2).value

    # Überprüfen, ob der Pfad vorhanden ist und ob er auf '.mp3' endet
    if path is not None and path.endswith('.mp3'): 
        source_file = os.path.join(source_folder_path, path)

        # Überprüfen, ob die Quelldatei existiert
        if os.path.isfile(source_file):
            destination_file = os.path.join(destination_folder_path, path)
            shutil.copy(source_file, destination_file) # Die Quelldatei zur Zieldatei kopieren
            print(f'Datei {path} kopiert.')